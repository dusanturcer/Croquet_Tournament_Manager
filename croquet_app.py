import streamlit as st
import pandas as pd
import sqlite3
import itertools
import random
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from datetime import datetime

# Database setup
def init_db():
    try:
        conn = sqlite3.connect('tournaments.db')
        c = conn.cursor()
        c.execute('DROP TABLE IF EXISTS tournaments')
        c.execute('''CREATE TABLE tournaments
                     (id INTEGER PRIMARY KEY, name TEXT, created_date TEXT,
                      players TEXT, num_rounds INTEGER, current_round INTEGER DEFAULT 1,
                      matches TEXT, standings TEXT, byes TEXT, pairing_method TEXT)''')
        conn.commit()
        conn.close()
        st.success("Database initialized successfully.")
        print("Database initialized successfully.")
    except sqlite3.OperationalError as e:
        st.error(f"Database initialization failed: {e}")
        print(f"Database initialization failed: {e}")
        st.stop()

def get_conn():
    return sqlite3.connect('tournaments.db')

# Helper functions
def sort_key(p):
    return (-p['score'], -p['net_hoops'], -p['hoops_scored'])

def generate_pairings(entities, pairing_method="Swiss", modifying=True):
    st.write(f"DEBUG: Generating pairings with method: {pairing_method}, Players: {[p['name'] for p in entities]}")
    print(f"DEBUG: Generating pairings with method: {pairing_method}, Players: {[p['name'] for p in entities]}")
    if not entities:
        st.error("No players provided for pairing.")
        print("ERROR: No players provided for pairing.")
        return [], [], False

    if len(entities) < 2:
        st.error("Need at least 2 players to generate pairings.")
        print("ERROR: Need at least 2 players to generate pairings.")
        return [], [], False

    if pairing_method == "Swiss":
        entity_list = sorted(entities, key=sort_key)
    else:  # Random
        entity_list = entities.copy()
        random.shuffle(entity_list)

    n = len(entity_list)
    best_pairings = []
    best_byes = []
    has_repeat = False
    min_repeats = float('inf')

    if n % 2 == 0:
        players_indices = list(range(n))
        possible_pairings = list(itertools.combinations(players_indices, 2))
        pairing_combinations = []
        for comb in itertools.combinations(possible_pairings, n // 2):
            players_covered = set()
            valid = True
            for p1, p2 in comb:
                if p1 in players_covered or p2 in players_covered:
                    valid = False
                    break
                players_covered.add(p1)
                players_covered.add(p2)
            if valid and len(players_covered) == n:
                pairing_combinations.append(comb)
    else:
        pairing_combinations = []
        for bye_idx in range(n):
            players_indices = [i for i in range(n) if i != bye_idx]
            possible_pairings = list(itertools.combinations(players_indices, 2))
            for comb in itertools.combinations(possible_pairings, (n - 1) // 2):
                players_covered = set()
                valid = True
                for p1, p2 in comb:
                    if p1 in players_covered or p2 in players_covered:
                        valid = False
                        break
                    players_covered.add(p1)
                    players_covered.add(p2)
                if valid and len(players_covered) == n - 1:
                    pairing_combinations.append((comb, bye_idx))

    if not pairing_combinations:
        st.error("No valid pairing combinations found.")
        print("ERROR: No valid pairing combinations found.")
        return [], [], False

    if pairing_method == "Random":
        random.shuffle(pairing_combinations)

    for comb in pairing_combinations:
        if n % 2 == 0:
            pairs = [(entity_list[p1]['name'], entity_list[p2]['name']) for p1, p2 in comb]
            bye = []
        else:
            pairs = [(entity_list[p1]['name'], entity_list[p2]['name']) for p1, p2 in comb[0]]
            bye = [entity_list[comb[1]]['name']]
        
        repeats = 0
        for p1, p2 in pairs:
            pl1 = next(p for p in entity_list if p['name'] == p1)
            if p2 in pl1['opponents']:
                repeats += 1
        if repeats < min_repeats:
            min_repeats = repeats
            best_pairings = pairs
            best_byes = bye
            has_repeat = repeats > 0
        if repeats == 0:
            break

    if modifying and best_pairings:
        for p1, p2 in best_pairings:
            pl1 = next(p for p in entity_list if p['name'] == p1)
            pl2 = next(p for p in entity_list if p['name'] == p2)
            pl1['opponents'].add(p2)
            pl2['opponents'].add(p1)

    st.write(f"DEBUG: Generated pairings: {best_pairings}, Byes: {best_byes}, Has repeat: {has_repeat}")
    print(f"DEBUG: Generated pairings: {best_pairings}, Byes: {best_byes}, Has repeat: {has_repeat}")
    return best_pairings, best_byes, has_repeat

def update_player_stats(pl, s_scored, s_conceded, is_win):
    pl['games_played'] += 1
    if is_win:
        pl['wins'] += 1
        pl['score'] += 1.0
    else:
        pl['losses'] += 1
    pl['hoops_scored'] += s_scored
    pl['hoops_conceded'] = s_conceded
    pl['net_hoops'] = pl['hoops_scored'] - pl['hoops_conceded']

def reset_player_stats(players):
    for p in players:
        p['score'] = 0.0
        p['games_played'] = 0
        p['wins'] = 0
        p['losses'] = 0
        p['hoops_scored'] = 0
        p['hoops_conceded'] = 0
        p['net_hoops'] = 0

# Initialize DB
try:
    init_db()
except sqlite3.OperationalError as e:
    st.error(f"Database initialization failed: {e}")
    print(f"Database initialization failed: {e}")
    st.stop()

# Streamlit App
st.markdown("<br>", unsafe_allow_html=True)
st.title("Croquet Tournament Manager")

# Sidebar
st.sidebar.title("Tournaments")

try:
    conn_temp = get_conn()
    tournament_list = pd.read_sql("SELECT id, name, created_date FROM tournaments", conn_temp)
    conn_temp.close()
    st.write(f"DEBUG: Loaded {len(tournament_list)} tournaments from database.")
    print(f"DEBUG: Loaded {len(tournament_list)} tournaments from database.")
except sqlite3.OperationalError as e:
    st.error(f"Failed to load tournaments: {e}")
    print(f"Failed to load tournaments: {e}")
    st.stop()

if 'selected_id' not in st.session_state:
    st.session_state.selected_id = 0

if not tournament_list.empty:
    options = [0] + list(tournament_list['id'])
    select_index = options.index(st.session_state.selected_id) if st.session_state.selected_id in options else 0
    selected_id = st.sidebar.selectbox(
        "Select Tournament:",
        options=options,
        format_func=lambda x: "New Tournament" if x == 0 else tournament_list[tournament_list['id'] == x]['name'].iloc[0],
        index=select_index,
        key="selectbox_tournament"
    )
    st.session_state.selected_id = selected_id
else:
    selected_id = 0
    st.session_state.selected_id = 0

if selected_id == 0:
    st.write("DEBUG: Rendering new_tournament form")
    print("DEBUG: Rendering new_tournament form")
    with st.form("new_tournament", clear_on_submit=True):
        tourney_name = st.text_input("Tournament Name:", key="tourney_name")
        num_players = st.number_input("Number of players:", min_value=2, value=4, key="num_players")
        num_rounds = st.number_input("Number of Rounds:", min_value=1, value=5, key="num_rounds")
        pairing_method = st.selectbox("Pairing Method:", ["Swiss", "Random"], key="pairing_method")
        submitted = st.form_submit_button("Next: Enter Player Names")
        if submitted:
            st.write(f"DEBUG: New tournament form submitted. Name: {tourney_name}, Players: {num_players}, Rounds: {num_rounds}, Method: {pairing_method}")
            print(f"DEBUG: New tournament form submitted. Name: {tourney_name}, Players: {num_players}, Rounds: {num_rounds}, Method: {pairing_method}")
            if not tourney_name:
                st.error("Please enter a tournament name.")
                print("ERROR: Tournament name empty.")
            else:
                st.session_state.num_players = num_players
                st.session_state.num_rounds = num_rounds
                st.session_state.tourney_name = tourney_name
                st.session_state.pairing_method = pairing_method
                st.session_state.form_step = "players"
                st.write(f"DEBUG: Session state set: {st.session_state}")
                print(f"DEBUG: Session state set: {st.session_state}")
    
    if st.session_state.get('form_step') == "players":
        st.write("DEBUG: Rendering players_form")
        print("DEBUG: Rendering players_form")
        with st.form("players_form", clear_on_submit=True):
            players = []
            all_names_filled = True
            unique_names = set()
            for i in range(st.session_state.num_players):
                name = st.text_input(f"Player {i+1} name:", key=f"player_{i}_{st.session_state.tourney_name}_{random.randint(1, 10000)}")
                if not name:
                    all_names_filled = False
                elif name in unique_names:
                    all_names_filled = False
                    st.error(f"Duplicate player name: {name}")
                    print(f"ERROR: Duplicate player name: {name}")
                else:
                    unique_names.add(name)
                    players.append({
                        'name': name, 'score': 0.0, 'games_played': 0, 'wins': 0, 'losses': 0,
                        'hoops_scored': 0, 'hoops_conceded': 0, 'net_hoops': 0, 'opponents': set()
                    })
            create_btn = st.form_submit_button("Create Tournament")
            if create_btn:
                st.write(f"DEBUG: Create Tournament clicked. All names filled: {all_names_filled}, Players: {[p['name'] for p in players if p['name']]}")
                print(f"DEBUG: Create Tournament clicked. All names filled: {all_names_filled}, Players: {[p['name'] for p in players if p['name']]}")
                if not all_names_filled:
                    st.error("Please fill all player names and ensure they are unique.")
                    print("ERROR: Not all player names filled or duplicates detected.")
                else:
                    try:
                        pairing_method = st.session_state.get('pairing_method', 'Swiss')
                        st.write(f"DEBUG: Creating tournament. Name: {st.session_state.tourney_name}, Pairing method: {pairing_method}")
                        print(f"DEBUG: Creating tournament. Name: {st.session_state.tourney_name}, Pairing method: {pairing_method}")
                        pairings, byes, has_repeat = generate_pairings(players, pairing_method)
                        if not pairings and not byes:
                            st.error("Failed to generate pairings. Please try again.")
                            print("ERROR: Failed to generate pairings.")
                            st.stop()
                        conn_temp = get_conn()
                        cur = conn_temp.cursor()
                        cur.execute(
                            "INSERT INTO tournaments (name, created_date, players, num_rounds, current_round, matches, standings, byes, pairing_method) VALUES (?, ?, ?, ?, 1, ?, ?, ?, ?)",
                            (st.session_state.tourney_name, datetime.now().isoformat(), str(players), st.session_state.num_rounds, str([]), str([]), str([byes]), pairing_method)
                        )
                        new_id = cur.lastrowid
                        conn_temp.commit()
                        conn_temp.close()
                        st.write(f"DEBUG: Tournament inserted with ID {new_id}")
                        print(f"DEBUG: Tournament inserted with ID {new_id}")
                        st.session_state.selected_id = new_id
                        st.session_state.current_pairings = pairings
                        st.session_state.current_byes = byes
                        st.session_state.has_repeat = has_repeat
                        st.session_state.current_round = 1
                        st.session_state.form_step = "created"
                        st.success(f"Tournament '{st.session_state.tourney_name}' created with ID {new_id}!")
                        print(f"DEBUG: Tournament '{st.session_state.tourney_name}' created with ID {new_id}")
                        for key in ['num_players', 'num_rounds', 'tourney_name', 'pairing_method', 'form_step']:
                            if key in st.session_state:
                                del st.session_state[key]
                        st.write("DEBUG: Session state cleared")
                        print("DEBUG: Session state cleared")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Failed to create tournament: {str(e)}")
                        print(f"ERROR: Failed to create tournament: {str(e)}")
                        st.stop()
else:
    try:
        conn_temp = get_conn()
        tourney_data = pd.read_sql("SELECT * FROM tournaments WHERE id=?", conn_temp, params=(selected_id,))
        conn_temp.close()
        st.write(f"DEBUG: Loaded tournament data for ID {selected_id}")
        print(f"DEBUG: Loaded tournament data for ID {selected_id}")
    except sqlite3.OperationalError as e:
        st.error(f"Failed to load tournament data: {e}")
        print(f"Failed to load tournament data: {e}")
        st.stop()

    if not tourney_data.empty:
        tourney = tourney_data.iloc[0].to_dict()
    else:
        st.error("Tournament not found!")
        print("ERROR: Tournament not found!")
        st.session_state.selected_id = 0
        st.rerun()
        st.stop()
    
    players = eval(tourney['players'])
    num_rounds = tourney['num_rounds']
    current_round = tourney['current_round']
    matches = eval(tourney['matches']) if tourney['matches'] else []
    standings_history = eval(tourney['standings']) if tourney['standings'] else []
    byes_history = eval(tourney['byes']) if tourney['byes'] else []
    pairing_method = tourney.get('pairing_method', 'Swiss')
    st.write(f"DEBUG: Loaded tournament. Name: {tourney['name']}, Pairing method: {pairing_method}")
    print(f"DEBUG: Loaded tournament. Name: {tourney['name']}, Pairing method: {pairing_method}")

    if current_round > num_rounds:
        st.header(f"Tournament: {tourney['name']} - Final Standings")
    else:
        st.header(f"Tournament: {tourney['name']} - Round {current_round} of {num_rounds}")

    # Current Standings
    st.subheader("Current Standings")
    if not standings_history:
        sorted_players = sorted(players, key=sort_key)
        current_standings = [
            {
                'rank': i + 1,
                'name': p['name'],
                'games_played': p['games_played'],
                'wins': p['wins'],
                'losses': p['losses'],
                'hoops_scored': p['hoops_scored'],
                'hoops_conceded': p['hoops_conceded'],
                'net_hoops': p['net_hoops'],
                'points': p['score'],
                'win_percentage': 0.00
            } for i, p in enumerate(sorted_players)
        ]
        df_stand = pd.DataFrame(current_standings)
    else:
        df_stand = pd.DataFrame(standings_history[-1])
        df_stand['win_percentage'] = (df_stand['wins'] / df_stand['games_played'] * 100).round(2).fillna(0.00)
    
    st.dataframe(df_stand, use_container_width=True, hide_index=True)

    if current_round <= num_rounds:
        if 'current_pairings' not in st.session_state or current_round != st.session_state.get('current_round', 0):
            pairings, byes, has_repeat = generate_pairings(players, pairing_method)
            st.session_state.current_pairings = pairings
            st.session_state.current_byes = byes
            st.session_state.has_repeat = has_repeat
            st.session_state.current_round = current_round
        else:
            pairings = st.session_state.current_pairings
            byes = st.session_state.current_byes
            has_repeat = st.session_state.has_repeat

        st.subheader(f"Round {current_round} Pairings")
        if has_repeat:
            st.warning("Some repeating pairings this round (unavoidable due to player count).")
        for i, (p1, p2) in enumerate(pairings, 1):
            st.write(f"{i}. {p1} vs {p2}")
        if byes:
            for b in byes:
                st.write(f"{b} gets a bye.")

        with st.form(f"results_round_{current_round}", clear_on_submit=True):
            result_data = {}
            for p1, p2 in pairings:
                col1, col2 = st.columns(2)
                with col1:
                    s1 = st.number_input(f"{p1} hoops:", min_value=0, key=f"s1_{p1}_{p2}_{current_round}_{random.randint(1, 10000)}")
                with col2:
                    s2 = st.number_input(f"{p2} hoops:", min_value=0, key=f"s2_{p1}_{p2}_{current_round}_{random.randint(1, 10000)}")
                result_data[(p1, p2)] = (s1, s2)
            submit_results = st.form_submit_button("Submit Results")

            if submit_results:
                new_matches = []
                for (p1, p2), (s1, s2) in result_data.items():
                    if s1 == 7 and s2 < 7:
                        is_win1 = True
                    elif s2 == 7 and s1 < 7:
                        is_win1 = False
                    else:
                        st.error("Invalid score: Must be first to 7.")
                        print("ERROR: Invalid score: Must be first to 7.")
                        st.stop()
                    
                    pl1 = next(p for p in players if p['name'] == p1)
                    pl2 = next(p for p in players if p['name'] == p2)
                    update_player_stats(pl1, s1, s2, is_win1)
                    update_player_stats(pl2, s2, s1, not is_win1)
                    new_matches.append({'round': current_round, 'player1': p1, 'player2': p2, 'score1': s1, 'score2': s2})
                
                for p1, p2 in pairings:
                    pl1 = next(p for p in players if p['name'] == p1)
                    pl2 = next(p for p in players if p['name'] == p2)
                    pl1['opponents'].add(p2)
                    pl2['opponents'].add(p1)
                
                matches.extend(new_matches)
                byes_history.append(byes)
                
                sorted_players = sorted(players, key=sort_key)
                standings_this = [
                    {
                        'rank': i + 1,
                        'name': p['name'],
                        'games_played': p['games_played'],
                        'wins': p['wins'],
                        'losses': p['losses'],
                        'hoops_scored': p['hoops_scored'],
                        'hoops_conceded': p['hoops_conceded'],
                        'net_hoops': p['net_hoops'],
                        'points': p['score'],
                        'win_percentage': (p['wins'] / p['games_played'] * 100) if p['games_played'] > 0 else 0.00
                    } for i, p in enumerate(sorted_players)
                ]
                standings_history.append(standings_this)
                
                try:
                    conn_temp = get_conn()
                    conn_temp.execute(
                        "UPDATE tournaments SET players=?, matches=?, standings=?, byes=?, current_round=? WHERE id=?",
                        (str(players), str(matches), str(standings_history), str(byes_history), current_round + 1, selected_id)
                    )
                    conn_temp.commit()
                    conn_temp.close()
                except sqlite3.OperationalError as e:
                    st.error(f"Failed to update tournament: {e}")
                    print(f"Failed to update tournament: {e}")
                    st.stop()
                
                if 'current_pairings' in st.session_state:
                    del st.session_state.current_pairings
                    del st.session_state.current_byes
                    del st.session_state.has_repeat
                    del st.session_state.current_round
                
                if current_round == num_rounds:
                    st.success("Tournament completed! Final standings updated.")
                    print("DEBUG: Tournament completed! Final standings updated.")
                else:
                    st.success("Results saved! Proceed to next round.")
                    print("DEBUG: Results saved! Proceed to next round.")
                st.rerun()

        if current_round < num_rounds and 'current_pairings' not in st.session_state:
            if st.button("Generate Next Round Pairings"):
                st.rerun()

    # Exports
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Export Matches CSV"):
            df_matches = pd.DataFrame(matches)
            csv = df_matches.to_csv(index=False)
            st.download_button("Download Matches", csv, "matches.csv", "text/csv")
    with col2:
        if st.button("Export Standings XLSX"):
            with pd.ExcelWriter('temp_standings.xlsx', engine='openpyxl') as writer:
                player_names = [p['name'] for p in players]
                sheet = "Final Standings"
                df_s = pd.DataFrame(standings_history[-1] if standings_history else [])
                if not df_s.empty:
                    df_s['win_percentage'] = (df_s['wins'] / df_stand['games_played'] * 100).round(2).fillna(0.00)
                else:
                    sorted_players = sorted(players, key=sort_key)
                    df_s = pd.DataFrame([
                        {
                            'rank': i + 1,
                            'name': p['name'],
                            'games_played': p['games_played'],
                            'wins': p['wins'],
                            'losses': p['losses'],
                            'hoops_scored': p['hoops_scored'],
                            'hoops_conceded': p['hoops_conceded'],
                            'net_hoops': p['net_hoops'],
                            'points': p['score'],
                            'win_percentage': 0.00
                        } for i, p in enumerate(sorted_players)
                    ])
                
                cross_table = pd.DataFrame(index=player_names, columns=player_names)
                cross_table.fillna('', inplace=True)
                for p in player_names:
                    cross_table.loc[p, p] = '-'
                for m in matches:
                    p1, p2 = m['player1'], m['player2']
                    s1, s2 = m['score1'], m['score2']
                    if s1 == 7 and s2 < 7:
                        cross_table.loc[p1, p2] = f"W {s1}-{s2}"
                        cross_table.loc[p2, p1] = f"L {s2}-{s1}"
                    elif s2 == 7 and s1 < 7:
                        cross_table.loc[p1, p2] = f"L {s1}-{s2}"
                        cross_table.loc[p2, p1] = f"W {s2}-{s1}"
                
                df_s.to_excel(writer, sheet, index=False, startrow=0)
                cross_table.to_excel(writer, sheet, index=True, startrow=len(df_s) + 2)
            
            wb = load_workbook('temp_standings.xlsx')
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center')
            wb.save('temp_standings.xlsx')
            with open('temp_standings.xlsx', 'rb') as f:
                st.download_button("Download Standings", f.read(), "standings.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Games Played
    if matches:
        st.header("Games Played")
        with st.form(f"edit_matches_form_{current_round}", clear_on_submit=True):
            edited_matches = []
            rounds = sorted(set(match['round'] for match in matches))
            for round_num in rounds:
                st.subheader(f"Round {round_num}")
                round_matches = [m for m in matches if m['round'] == round_num]
                for idx, match in enumerate(round_matches):
                    col1, col2, col3, col4 = st.columns([1, 2, 1, 2])
                    with col1:
                        st.write(match['player1'])
                    with col2:
                        s1 = st.number_input(
                            f"{match['player1']} score",
                            min_value=0,
                            value=match['score1'],
                            key=f"edit_s1_{idx}_{match['round']}_{match['player1']}_{match['player2']}_{random.randint(1, 10000)}"
                        )
                    with col3:
                        st.write(match['player2'])
                    with col4:
                        s2 = st.number_input(
                            f"{match['player2']} score",
                            min_value=0,
                            value=match['score2'],
                            key=f"edit_s2_{idx}_{match['round']}_{match['player1']}_{match['player2']}_{random.randint(1, 10000)}"
                        )
                    edited_matches.append({
                        'round': match['round'],
                        'player1': match['player1'],
                        'player2': match['player2'],
                        'score1': s1,
                        'score2': s2
                    })
            
            update_standings = st.form_submit_button("Update Standings")
            
            if update_standings:
                for match in edited_matches:
                    if (match['score1'] == 7 and match['score2'] < 7) or (match['score2'] == 7 and match['score1'] < 7):
                        continue
                    else:
                        st.error(f"Invalid score in Round {match['round']} for {match['player1']} vs {match['player2']}: Must be first to 7.")
                        print(f"ERROR: Invalid score in Round {match['round']} for {match['player1']} vs {match['player2']}: Must be first to 7.")
                        st.stop()
                
                reset_player_stats(players)
                
                for match in edited_matches:
                    pl1 = next(p for p in players if p['name'] == match['player1'])
                    pl2 = next(p for p in players if p['name'] == match['player2'])
                    is_win1 = match['score1'] == 7 and match['score2'] < 7
                    update_player_stats(pl1, match['score1'], match['score2'], is_win1)
                    update_player_stats(pl2, match['score2'], match['score1'], not is_win1)
                
                sorted_players = sorted(players, key=sort_key)
                standings_this = [
                    {
                        'rank': i + 1,
                        'name': p['name'],
                        'games_played': p['games_played'],
                        'wins': p['wins'],
                        'losses': p['losses'],
                        'hoops_scored': p['hoops_scored'],
                        'hoops_conceded': p['hoops_conceded'],
                        'net_hoops': p['net_hoops'],
                        'points': p['score'],
                        'win_percentage': (p['wins'] / p['games_played'] * 100) if p['games_played'] > 0 else 0.00
                    } for i, p in enumerate(sorted_players)
                ]
                standings_history.append(standings_this)
                
                try:
                    conn_temp = get_conn()
                    conn_temp.execute(
                        "UPDATE tournaments SET players=?, matches=?, standings=? WHERE id=?",
                        (str(players), str(edited_matches), str(standings_history), selected_id)
                    )
                    conn_temp.commit()
                    conn_temp.close()
                except sqlite3.OperationalError as e:
                    st.error(f"Failed to update standings: {e}")
                    print(f"Failed to update standings: {e}")
                    st.stop()
                
                st.success("Standings updated based on edited match results!")
                print("DEBUG: Standings updated based on edited match results!")
                st.rerun()

if selected_id != 0:
    if st.sidebar.button("Delete Tournament"):
        try:
            conn_temp = get_conn()
            conn_temp.execute("DELETE FROM tournaments WHERE id=?", (selected_id,))
            conn_temp.commit()
            conn_temp.close()
            st.session_state.selected_id = 0
            if 'current_pairings' in st.session_state:
                del st.session_state.current_pairings
                del st.session_state.current_byes
                del st.session_state.has_repeat
                del st.session_state.current_round
            st.sidebar.success("Tournament deleted!")
            print("DEBUG: Tournament deleted!")
            st.rerun()
        except sqlite3.OperationalError as e:
            st.error(f"Failed to delete tournament: {e}")
            print(f"Failed to delete tournament: {e}")
            st.stop()